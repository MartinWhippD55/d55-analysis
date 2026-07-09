"""Spreadsheet generators (Task 12).

Two generators, both driven by the programme layout so outputs land on the
correct side of the internal/client split:

- :func:`generate_delivery_playbook` — the **internal** Delivery Playbook runbook
  (stages, activities, owners, inputs/outputs, decision points). Written under
  ``internal/`` and never handed to a client (Requirement 9.1, 9.2).
- :func:`generate_questionnaire_spreadsheet` — the **client-facing** assessment
  questionnaire (questions per dimension + the 1–5 scale). Written under
  ``client/`` with dimension names taken from the manifest so they match exactly
  (Requirement 10.1, 10.2).

Uses ``openpyxl``. Save paths come from the layout — never a hard-coded absolute
path.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Mapping, Sequence

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .layout import ProgrammeLayout
from .manifest import dimension_names, load_manifest

# Shared styling.
_TITLE_FONT = Font(name="Calibri", size=16, bold=True)
_SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="64748B")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _title(ws, title: str, subtitle: str, span: int) -> None:
    last_col = get_column_letter(span)
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].font = _SUBTITLE_FONT


def _header_row(ws, columns: Sequence[str], row: int) -> None:
    for i, col in enumerate(columns, 1):
        cell = ws.cell(row=row, column=i, value=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _write_rows(ws, rows: Sequence[Sequence[Any]], start_row: int) -> int:
    r = start_row
    for row in rows:
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=value)
            cell.alignment = _WRAP
        r += 1
    return r


# ---------------------------------------------------------------------------
# 12.1 — internal Delivery Playbook (written under internal/)
# ---------------------------------------------------------------------------

DELIVERY_PLAYBOOK_COLUMNS = ("Stage", "Activity", "Owner", "Inputs", "Outputs", "Decision point")

# A sensible default runbook keyed to the four programme stages.
_DEFAULT_ACTIVITIES: tuple[dict[str, str], ...] = (
    {"stage": "Scope & Frame", "activity": "Establish programme context and dimensions",
     "owner": "Lead consultant", "inputs": "Client intake, assessment scores",
     "outputs": "context.md, dimensions.md", "decision": "Scope agreed at human gate?"},
    {"stage": "Build Modules", "activity": "Author in-scope modules to schema",
     "owner": "Consultant", "inputs": "Recommended modules, dimensions",
     "outputs": "module.md per module", "decision": "Module content approved?"},
    {"stage": "Generate Assets", "activity": "Render per-module deliverables + spreadsheets",
     "owner": "Consultant", "inputs": "Approved modules, brand config",
     "outputs": "HTML/PDF assets, questionnaire", "decision": "Assets verified?"},
    {"stage": "Verify & Ship", "activity": "Verify outputs and assemble client bundle",
     "owner": "Lead consultant", "inputs": "All generated assets",
     "outputs": "Client bundle (excludes internal/)", "decision": "Ready to ship?"},
)


def generate_delivery_playbook(
    layout: ProgrammeLayout,
    activities: Sequence[Mapping[str, str]] | None = None,
    programme_name: str | None = None,
    filename: str = "Delivery Playbook.xlsx",
) -> Path:
    """Generate the internal Delivery Playbook runbook under ``internal/``."""
    if programme_name is None:
        programme_name = load_manifest(layout.root).get("programme", {}).get("name", "Programme")
    activities = activities or _DEFAULT_ACTIVITIES

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Delivery Playbook"
    for i, w in enumerate((20, 40, 18, 32, 32, 30), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title(ws, f"{programme_name} — Delivery Playbook", "INTERNAL — D55 delivery runbook. Not for client distribution.", len(DELIVERY_PLAYBOOK_COLUMNS))
    _header_row(ws, DELIVERY_PLAYBOOK_COLUMNS, 4)
    rows = [
        (a["stage"], a["activity"], a["owner"], a.get("inputs", ""), a.get("outputs", ""), a.get("decision", ""))
        for a in activities
    ]
    _write_rows(ws, rows, 5)

    layout.internal_dir.mkdir(parents=True, exist_ok=True)
    out = layout.internal_dir / filename
    wb.save(out)
    return out


# ---------------------------------------------------------------------------
# 12.2 — client-facing assessment questionnaire (written under client/)
# ---------------------------------------------------------------------------

QUESTIONNAIRE_COLUMNS = ("Dimension", "Tier", "Question", "Current (1-5)", "Target (1-5)", "Notes")


def generate_questionnaire_spreadsheet(
    layout: ProgrammeLayout,
    questions: Mapping[str, Mapping[str, Sequence[str]]] | None = None,
    programme_name: str | None = None,
    filename: str = "assessment-questionnaire.xlsx",
) -> Path:
    """Generate the client-facing assessment questionnaire under ``client/``.

    ``questions`` maps a dimension name -> ``{"must_ask": [...], "go_deeper": [...]}``.
    Dimension names are taken from the manifest so they match exactly
    (Requirement 10.2); a dimension with no supplied questions still gets a row.
    """
    manifest = load_manifest(layout.root)
    if programme_name is None:
        programme_name = manifest.get("programme", {}).get("name", "Programme")
    dims = dimension_names(manifest)
    questions = questions or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assessment Questionnaire"
    for i, w in enumerate((32, 12, 60, 14, 14, 30), 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    _title(ws, f"{programme_name} — Assessment Questionnaire",
           "Score each dimension 1–5: current vs target. 1 = absent, 5 = strategic/advanced.",
           len(QUESTIONNAIRE_COLUMNS))
    _header_row(ws, QUESTIONNAIRE_COLUMNS, 4)

    rows: list[tuple] = []
    for dim in dims:
        dq = questions.get(dim, {})
        must = dq.get("must_ask") or []
        deep = dq.get("go_deeper") or []
        if not must and not deep:
            rows.append((dim, "", "TODO — add questions for this dimension", "", "", ""))
            continue
        for q in must:
            rows.append((dim, "Must-ask", q, "", "", ""))
        for q in deep:
            rows.append((dim, "Go deeper", q, "", "", ""))
    _write_rows(ws, rows, 5)

    # A second sheet documenting the 1–5 scale.
    scale = wb.create_sheet("Scoring Scale")
    scale.column_dimensions["A"].width = 10
    scale.column_dimensions["B"].width = 60
    _header_row(scale, ("Score", "Meaning"), 1)
    _write_rows(scale, [
        (1, "Absent — no capability in place."),
        (2, "Aware — informal / ad-hoc."),
        (3, "Endorsed — provisioned but inconsistent."),
        (4, "Effective — embedded for the majority."),
        (5, "Strategic — a stated competitive advantage."),
    ], 2)

    layout.client_dir.mkdir(parents=True, exist_ok=True)
    out = layout.client_dir / filename
    wb.save(out)
    return out


__all__ = [
    "DELIVERY_PLAYBOOK_COLUMNS",
    "QUESTIONNAIRE_COLUMNS",
    "generate_delivery_playbook",
    "generate_questionnaire_spreadsheet",
]
