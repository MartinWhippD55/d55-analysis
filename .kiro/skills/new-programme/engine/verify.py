"""Output verification harness (Task 15, Requirement 15).

Verifies produced assets before a programme is declared ready and returns
structured reports (rather than raising) so a caller can regenerate the specific
failing asset and re-verify:

- :func:`verify_html` — drives Playwright to measure the DOM: images loaded,
  expected element counts, and no horizontal overflow (15.1).
- :func:`verify_pdf` — reads the PDF with ``pypdf`` for page count, page size
  (A4), and a best-effort orphaned-heading check (15.2).
- :func:`verify_xlsx` — opens the workbook with ``openpyxl`` and checks expected
  sheets and row counts (15.3).

Temp resources (the optional local server) are cleaned up via context managers
(15.5). HTML is served over a short-lived localhost server so verification
matches how a browser would load it; self-contained files also work over
``file://`` when ``serve=False``.
"""
from __future__ import annotations

import contextlib
import functools
import http.server
import socketserver
import threading
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator, Mapping, Sequence


# ---------------------------------------------------------------------------
# Report types
# ---------------------------------------------------------------------------

@dataclass
class VerifyIssue:
    asset: str
    kind: str
    detail: str


@dataclass
class VerifyReport:
    asset: str
    ok: bool
    issues: list[VerifyIssue] = field(default_factory=list)
    metrics: dict = field(default_factory=dict)

    def _add(self, kind: str, detail: str) -> None:
        self.issues.append(VerifyIssue(self.asset, kind, detail))
        self.ok = False


# ---------------------------------------------------------------------------
# Local static server (15.5 — cleaned up on exit)
# ---------------------------------------------------------------------------

@contextlib.contextmanager
def serve_directory(directory: Path) -> Iterator[str]:
    """Serve ``directory`` on an ephemeral localhost port; shut down on exit."""
    handler = functools.partial(http.server.SimpleHTTPRequestHandler, directory=str(directory))
    with socketserver.TCPServer(("127.0.0.1", 0), handler) as httpd:
        port = httpd.server_address[1]
        thread = threading.Thread(target=httpd.serve_forever, daemon=True)
        thread.start()
        try:
            yield f"http://127.0.0.1:{port}"
        finally:
            httpd.shutdown()
            thread.join(timeout=5)


# ---------------------------------------------------------------------------
# HTML verification (15.1)
# ---------------------------------------------------------------------------

def verify_html(
    html_path: Path,
    *,
    expect_selectors: Mapping[str, int] | None = None,
    require_images: bool = True,
    viewport_width: int = 1280,
    serve: bool = True,
) -> VerifyReport:
    """Measure a rendered HTML file's DOM and report issues.

    ``expect_selectors`` maps a CSS selector -> minimum expected count. Fails if
    any image did not load, any selector is under its minimum, or the page
    overflows horizontally.
    """
    from playwright.sync_api import sync_playwright

    html_path = Path(html_path)
    report = VerifyReport(asset=str(html_path), ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page(viewport={"width": viewport_width, "height": 900})
        try:
            if serve:
                with serve_directory(html_path.parent) as base:
                    page.goto(f"{base}/{html_path.name}", wait_until="load")
                    _measure_html(page, report, expect_selectors, require_images, viewport_width)
            else:
                page.goto(html_path.resolve().as_uri(), wait_until="load")
                _measure_html(page, report, expect_selectors, require_images, viewport_width)
        finally:
            browser.close()
    return report


def _measure_html(page, report, expect_selectors, require_images, viewport_width) -> None:
    images = page.evaluate(
        "() => { const i = Array.from(document.images);"
        " return { total: i.length, loaded: i.filter(x => x.complete && x.naturalWidth > 0).length }; }"
    )
    report.metrics["images"] = images
    if require_images and images["total"] and images["loaded"] < images["total"]:
        report._add("images_not_loaded", f"{images['loaded']}/{images['total']} images loaded")

    overflow = page.evaluate(
        "() => document.documentElement.scrollWidth > document.documentElement.clientWidth + 1"
    )
    report.metrics["horizontal_overflow"] = overflow
    if overflow:
        sw = page.evaluate("() => document.documentElement.scrollWidth")
        report._add("horizontal_overflow", f"scrollWidth {sw} > viewport {viewport_width}")

    counts: dict[str, int] = {}
    for selector, minimum in (expect_selectors or {}).items():
        n = page.evaluate("(s) => document.querySelectorAll(s).length", selector)
        counts[selector] = n
        if n < minimum:
            report._add("selector_count", f"{selector!r}: {n} < expected {minimum}")
    report.metrics["selectors"] = counts


# ---------------------------------------------------------------------------
# PDF verification (15.2)
# ---------------------------------------------------------------------------

_A4_MM = (210.0, 297.0)


def verify_pdf(
    pdf_path: Path,
    *,
    expected_pages: int | None = None,
    page_size: str = "A4",
    heading_texts: Sequence[str] | None = None,
) -> VerifyReport:
    """Verify a PDF's page count, page size, and (best-effort) orphaned headings.

    ``heading_texts`` (e.g. the document's H2/H3 strings) enables orphan
    detection: a heading that is the last text on a page (with content on the
    next) is flagged. Without it, the orphan check is skipped.
    """
    from pypdf import PdfReader

    pdf_path = Path(pdf_path)
    report = VerifyReport(asset=str(pdf_path), ok=True)
    reader = PdfReader(str(pdf_path))
    pages = len(reader.pages)
    report.metrics["pages"] = pages

    if expected_pages is not None and pages != expected_pages:
        report._add("page_count", f"{pages} pages, expected {expected_pages}")

    if page_size.upper() == "A4":
        box = reader.pages[0].mediabox
        w_mm = float(box.width) / 72 * 25.4
        h_mm = float(box.height) / 72 * 25.4
        report.metrics["page_size_mm"] = (round(w_mm, 1), round(h_mm, 1))
        if abs(w_mm - _A4_MM[0]) > 3 or abs(h_mm - _A4_MM[1]) > 3:
            report._add("page_size", f"{w_mm:.0f}x{h_mm:.0f}mm, expected A4 210x297mm")

    if heading_texts:
        headings = {h.strip() for h in heading_texts}
        for idx in range(pages - 1):  # last page can legitimately end on a heading
            text = reader.pages[idx].extract_text() or ""
            lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
            if lines and lines[-1] in headings:
                report._add("orphaned_heading", f"page {idx + 1} ends with heading {lines[-1]!r}")

    return report


# ---------------------------------------------------------------------------
# Spreadsheet verification (15.3)
# ---------------------------------------------------------------------------

def verify_xlsx(
    xlsx_path: Path,
    *,
    expected_sheets: Sequence[str] | None = None,
    min_rows: Mapping[str, int] | None = None,
) -> VerifyReport:
    """Verify a workbook has the expected sheets and minimum row counts."""
    import openpyxl

    xlsx_path = Path(xlsx_path)
    report = VerifyReport(asset=str(xlsx_path), ok=True)
    wb = openpyxl.load_workbook(xlsx_path)
    report.metrics["sheets"] = list(wb.sheetnames)

    for sheet in expected_sheets or ():
        if sheet not in wb.sheetnames:
            report._add("missing_sheet", f"expected sheet {sheet!r}")

    rows_metric: dict[str, int] = {}
    for sheet, minimum in (min_rows or {}).items():
        if sheet not in wb.sheetnames:
            report._add("missing_sheet", f"expected sheet {sheet!r}")
            continue
        n = wb[sheet].max_row
        rows_metric[sheet] = n
        if n < minimum:
            report._add("row_count", f"{sheet!r}: {n} rows < expected {minimum}")
    report.metrics["rows"] = rows_metric
    return report


__all__ = [
    "VerifyIssue",
    "VerifyReport",
    "serve_directory",
    "verify_html",
    "verify_pdf",
    "verify_xlsx",
]
