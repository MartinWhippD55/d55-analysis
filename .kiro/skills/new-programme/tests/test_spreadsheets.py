"""Tests for the spreadsheet generators (Task 12)."""
from __future__ import annotations

import openpyxl

from engine.layout import template_layout
from engine.manifest import write_manifest
from engine.spreadsheets import (
    DELIVERY_PLAYBOOK_COLUMNS,
    QUESTIONNAIRE_COLUMNS,
    generate_delivery_playbook,
    generate_questionnaire_spreadsheet,
)

DIMENSIONS = ["Leadership & Mandate", "Metrics & ROI", "Governance, Security & Compliance"]


def _layout_with_manifest(tmp_path):
    layout = template_layout(tmp_path, "demo").create()
    write_manifest(
        {
            "programme": {"slug": "demo", "name": "Demo Programme"},
            "dimensions": [{"name": d} for d in DIMENSIONS],
            "modules": [],
        },
        layout.root,
    )
    return layout


def _header(ws, ncols):
    return [ws.cell(row=4, column=i).value for i in range(1, ncols + 1)]


# ---------------------------------------------------------------------------
# 12.1 — Delivery Playbook (internal)
# ---------------------------------------------------------------------------

def test_delivery_playbook_written_under_internal(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    out = generate_delivery_playbook(layout)
    assert out.parent == layout.internal_dir
    assert out.exists()
    # Never under client/.
    assert not (layout.client_dir / out.name).exists()


def test_delivery_playbook_has_expected_columns_and_rows(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    out = generate_delivery_playbook(layout)
    ws = openpyxl.load_workbook(out)["Delivery Playbook"]
    assert _header(ws, len(DELIVERY_PLAYBOOK_COLUMNS)) == list(DELIVERY_PLAYBOOK_COLUMNS)
    # Default runbook has the four programme stages.
    stages = {ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)}
    assert {"Scope & Frame", "Build Modules", "Generate Assets", "Verify & Ship"} <= stages


def test_delivery_playbook_accepts_custom_activities(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    activities = [{"stage": "Kickoff", "activity": "Do X", "owner": "Me",
                   "inputs": "A", "outputs": "B", "decision": "Go?"}]
    out = generate_delivery_playbook(layout, activities=activities)
    ws = openpyxl.load_workbook(out)["Delivery Playbook"]
    assert ws.cell(row=5, column=1).value == "Kickoff"


# ---------------------------------------------------------------------------
# 12.2 — Assessment questionnaire (client-facing)
# ---------------------------------------------------------------------------

def test_questionnaire_written_under_client(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    out = generate_questionnaire_spreadsheet(layout)
    assert out.parent == layout.client_dir
    assert out.exists()
    assert not (layout.internal_dir / out.name).exists()


def test_questionnaire_dimension_names_match_manifest_exactly(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    out = generate_questionnaire_spreadsheet(layout)
    ws = openpyxl.load_workbook(out)["Assessment Questionnaire"]
    assert _header(ws, len(QUESTIONNAIRE_COLUMNS)) == list(QUESTIONNAIRE_COLUMNS)
    seen = {ws.cell(row=r, column=1).value for r in range(5, ws.max_row + 1)}
    assert seen == set(DIMENSIONS)               # exact match, no drift


def test_questionnaire_includes_questions_and_scale(tmp_path):
    layout = _layout_with_manifest(tmp_path)
    questions = {
        "Leadership & Mandate": {
            "must_ask": ["Who sponsors AI tooling?"],
            "go_deeper": ["Who fights for the budget if cut?"],
        }
    }
    out = generate_questionnaire_spreadsheet(layout, questions=questions)
    wb = openpyxl.load_workbook(out)
    ws = wb["Assessment Questionnaire"]
    rows = [tuple(ws.cell(row=r, column=c).value for c in range(1, 4)) for r in range(5, ws.max_row + 1)]
    assert ("Leadership & Mandate", "Must-ask", "Who sponsors AI tooling?") in rows
    assert ("Leadership & Mandate", "Go deeper", "Who fights for the budget if cut?") in rows
    # Scale sheet documents 1–5.
    assert "Scoring Scale" in wb.sheetnames
    scale = wb["Scoring Scale"]
    scores = {scale.cell(row=r, column=1).value for r in range(2, 7)}
    assert scores == {1, 2, 3, 4, 5}
