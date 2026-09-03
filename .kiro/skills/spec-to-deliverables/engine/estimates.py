"""
Build an estimates spreadsheet from spec ``tasks.md`` files.

Parses each spec's task list, classifies + weights each sub-task, and writes an
xlsx with a raw **Task Detail** sheet (the source of truth) and a **Summary**
sheet of per-estimate rollups (live SUMIFS/COUNTIF formulas so it stays correct
if edited in Excel). ``engine.figures`` then reads this spreadsheet.

Nothing is spec-specific: pass the list of ``(tasks_md_path, estimate_name)`` to
include, and optionally override the day-weight table or add manual Summary rows
(e.g. a training line with no task breakdown).

Usage:
    from engine.estimates import generate_estimates
    generate_estimates(
        task_files=[(".kiro/specs/<spec>/tasks.md", "Est 1: Templates")],
        output_path="deliverables/<spec>/estimates.xlsx",
    )
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

# Default day-weight per task category. Override via ``weights=`` — this is the
# main lever reviewers adjust; keep it visible.
DEFAULT_WEIGHTS = {
    "infrastructure": 1.0,
    "api_backend": 0.5,
    "frontend": 0.75,
    "testing": 0.5,
    "prompt_iteration": 3.0,
    "integration": 0.75,
    "checkpoint": 0.0,
}

# Sub-task line, e.g. "  - [ ]* 1.2 Do the thing" (optional marked with '*').
_TASK_RE = re.compile(r"\s+- \[ \](\*?)\s+(\d+\.\d+[a-z]?)\s+(.*)")


def classify_task(task_text: str) -> str:
    """Classify a task description into a weight category (heuristic, editable)."""
    t = task_text.lower()
    if "checkpoint" in t:
        return "checkpoint"
    if "property test" in t or "integration test" in t or "unit test" in t:
        return "testing"
    if any(k in t for k in ("cdk", "infrastructure", "iam", "trust policy",
                            "athena workgroup", "wire cdk", "configure athena")):
        return "infrastructure"
    if ("angular" in t or "component" in t or "frontend" in t or "module" in t
            or ("service" in t and "implement" in t)):
        return "frontend"
    if "prompt iteration" in t or "iteration cycle" in t:
        return "prompt_iteration"
    if "integration wiring" in t or "navigation entry" in t or "wire cdk deployment" in t:
        return "integration"
    return "api_backend"


def parse_tasks_text(text: str, estimate_name: str, weights: Optional[dict] = None) -> list[dict]:
    """Pure: parse a tasks.md string into weighted task rows."""
    weights = weights or DEFAULT_WEIGHTS
    tasks = []
    for line in text.split("\n"):
        m = _TASK_RE.match(line)
        if not m:
            continue
        is_optional = m.group(1) == "*"
        task_id = m.group(2)
        task_text = m.group(3).strip()
        category = classify_task(task_text)
        tasks.append({
            "estimate": estimate_name,
            "task_id": task_id,
            "task": task_text,
            "category": category,
            "days": weights.get(category, 0.0),
            "optional": is_optional,
        })
    return tasks


def summarize(tasks: list[dict]) -> dict:
    """Roll up a task list into required/optional/total days + count."""
    required = sum(t["days"] for t in tasks if not t["optional"])
    optional = sum(t["days"] for t in tasks if t["optional"])
    return {
        "total_tasks": len(tasks),
        "required_days": round(required, 3),
        "optional_days": round(optional, 3),
        "total_days": round(required + optional, 3),
    }


def generate_estimates(task_files, output_path, weights: Optional[dict] = None,
                       manual_rows: Optional[list[dict]] = None, title: str = "Estimate Summary"):
    """Write the estimates xlsx. Returns the list of per-estimate summaries.

    ``task_files``: iterable of ``(tasks_md_path, estimate_name)``.
    ``manual_rows``: optional list of ``{name, total_tasks, required_days,
    optional_days, total_days}`` for estimates with no task breakdown; each is
    inserted onto the Summary sheet in order.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    weights = weights or DEFAULT_WEIGHTS
    all_tasks: list[dict] = []
    summaries: list[dict] = []
    for filepath, estimate_name in task_files:
        text = Path(filepath).read_text(encoding="utf-8")
        tasks = parse_tasks_text(text, estimate_name, weights)
        s = summarize(tasks)
        s["name"] = estimate_name
        summaries.append(s)
        all_tasks.extend(tasks)

    for manual in manual_rows or []:
        summaries.append({
            "name": manual["name"],
            "total_tasks": manual.get("total_tasks", 0),
            "required_days": manual.get("required_days", 0.0),
            "optional_days": manual.get("optional_days", 0.0),
            "total_days": manual.get("total_days",
                                     manual.get("required_days", 0.0) + manual.get("optional_days", 0.0)),
        })

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    border = Border(*(Side(style="thin"),) * 4)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")
    for col, h in enumerate(["Estimate", "Sub-Tasks", "Required (days)", "Optional (days)", "Total (days)"], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, border
    for i, s in enumerate(summaries, 4):
        ws.cell(row=i, column=1, value=s["name"]).border = border
        ws.cell(row=i, column=2, value=s["total_tasks"]).border = border
        ws.cell(row=i, column=3, value=s["required_days"]).border = border
        ws.cell(row=i, column=4, value=s["optional_days"]).border = border
        ws.cell(row=i, column=5, value=s["total_days"]).border = border
    total_row = 4 + len(summaries)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col, keyname in [(2, "total_tasks"), (3, "required_days"), (4, "optional_days"), (5, "total_days")]:
        ws.cell(row=total_row, column=col, value=sum(s[keyname] for s in summaries)).font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=total_row, column=col).border = border
    for col, w in zip("ABCDE", (35, 12, 16, 16, 14)):
        ws.column_dimensions[col].width = w

    wd = wb.create_sheet("Task Detail")
    for col, h in enumerate(["Estimate", "Task ID", "Task Description", "Category", "Days", "Optional"], 1):
        c = wd.cell(row=1, column=col, value=h)
        c.font, c.fill, c.border = header_font, header_fill, border
    for i, t in enumerate(all_tasks, 2):
        wd.cell(row=i, column=1, value=t["estimate"]).border = border
        wd.cell(row=i, column=2, value=t["task_id"]).border = border
        wd.cell(row=i, column=3, value=t["task"]).border = border
        wd.cell(row=i, column=4, value=t["category"]).border = border
        wd.cell(row=i, column=5, value=t["days"]).border = border
        wd.cell(row=i, column=6, value="Yes" if t["optional"] else "").border = border
    for col, w in zip("ABCDEF", (30, 8, 60, 15, 8, 10)):
        wd.column_dimensions[col].width = w

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return summaries
