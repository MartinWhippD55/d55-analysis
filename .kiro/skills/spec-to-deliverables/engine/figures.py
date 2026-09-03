"""
Single source of truth for estimate figures.

Reads a per-spec estimates spreadsheet (produced by ``engine.estimates``) and
computes per-estimate day figures directly from the raw "Task Detail" rows — the
same logic the Summary sheet's SUMIFS formulas use. Every generator imports from
here rather than hardcoding numbers, so a single spreadsheet edit + regenerate
propagates everywhere.

Nothing is spec-specific: estimate keys are derived from the estimate names in
the sheet (or an explicit ``name_to_key`` map for stable short keys like
``est1``). Estimates with no task breakdown (e.g. a training line) are read from
the Summary sheet as manual rows.

Usage:
    from engine.figures import load_figures, fmt
    figs = load_figures("deliverables/<spec>/estimates.xlsx",
                        name_to_key={"Est 1: Templates": "est1"})
    figs.get("est1").total      # -> 13.5
    fmt(figs.get("est1").total) # -> "13.5"
    figs.effort_line("est1")    # -> "~13.5 developer days (9.0 required + testing)"
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional


@dataclass(frozen=True)
class Figure:
    key: str
    name: str
    required: float
    optional: float
    total: float
    task_count: int


def fmt(value: float) -> str:
    """Format a day figure: one decimal place, trailing '.0' trimmed to whole days."""
    rounded = round(float(value), 1)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:.1f}"


def _slug_key(name: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", str(name).lower()).strip("-")
    return s or "est"


@dataclass
class Figures:
    """A collection of estimate figures, indexed by both stable key and name."""

    by_key: dict[str, Figure] = field(default_factory=dict)
    by_name: dict[str, Figure] = field(default_factory=dict)

    def get(self, key_or_name: str) -> Figure:
        if key_or_name in self.by_key:
            return self.by_key[key_or_name]
        if key_or_name in self.by_name:
            return self.by_name[key_or_name]
        raise KeyError(key_or_name)

    def keys(self):
        return list(self.by_key.keys())

    def effort_line(self, key_or_name: str) -> str:
        f = self.get(key_or_name)
        return f"~{fmt(f.total)} developer days ({fmt(f.required)} required + testing)"

    def grand_total(self, keys: Optional[list[str]] = None) -> Figure:
        figs = [self.get(k) for k in keys] if keys else list(self.by_key.values())
        req = sum(f.required for f in figs)
        opt = sum(f.optional for f in figs)
        count = sum(f.task_count for f in figs)
        return Figure("total", "TOTAL", round(req, 3), round(opt, 3), round(req + opt, 3), count)


def compute_figures(detail_rows, summary_rows=None, name_to_key: Optional[dict] = None) -> Figures:
    """Pure core: build Figures from raw rows (no file I/O).

    ``detail_rows``: iterable of ``(estimate_name, days, optional_flag)`` where
    ``optional_flag`` is truthy / "Yes" for optional tasks.
    ``summary_rows``: optional iterable of ``(name, count, required, optional, total)``
    for estimates that have no task breakdown (manual Summary rows). Only names not
    already present from detail rows are added.
    """
    name_to_key = name_to_key or {}
    acc: dict[str, dict] = {}
    order: list[str] = []
    for name, days, optional in detail_rows:
        if not name or days is None:
            continue
        if name not in acc:
            acc[name] = {"required": 0.0, "optional": 0.0, "count": 0}
            order.append(name)
        acc[name]["count"] += 1
        is_optional = optional in (True, "Yes", "yes", "Y", "y")
        if is_optional:
            acc[name]["optional"] += float(days)
        else:
            acc[name]["required"] += float(days)

    figs = Figures()
    for name in order:
        e = acc[name]
        key = name_to_key.get(name, _slug_key(name))
        f = Figure(key, name, round(e["required"], 3), round(e["optional"], 3),
                   round(e["required"] + e["optional"], 3), e["count"])
        figs.by_key[key] = f
        figs.by_name[name] = f

    for name, count, required, optional, total in summary_rows or []:
        if not name or name in acc:
            continue
        if str(name).strip().upper() == "TOTAL":   # totals row is not an estimate
            continue
        if not isinstance(required, (int, float)) or not isinstance(optional, (int, float)):
            continue
        key = name_to_key.get(name, _slug_key(name))
        t = float(total) if isinstance(total, (int, float)) else float(required) + float(optional)
        f = Figure(key, name, float(required), float(optional), round(t, 3),
                   int(count) if isinstance(count, (int, float)) else 0)
        figs.by_key[key] = f
        figs.by_name[name] = f

    return figs


def load_figures(spreadsheet_path, name_to_key: Optional[dict] = None,
                 detail_sheet: str = "Task Detail", summary_sheet: str = "Summary") -> Figures:
    """Load figures from an xlsx. Reads raw Task Detail rows (not cached formulas)."""
    from openpyxl import load_workbook

    path = Path(spreadsheet_path)
    if not path.exists():
        raise FileNotFoundError(f"Estimates spreadsheet not found: {path}")
    wb = load_workbook(path, data_only=False)

    detail = wb[detail_sheet]
    detail_rows = []
    for row in range(2, detail.max_row + 1):
        detail_rows.append((
            detail.cell(row=row, column=1).value,   # estimate name
            detail.cell(row=row, column=5).value,    # days
            detail.cell(row=row, column=6).value,    # optional flag
        ))

    summary_rows = []
    if summary_sheet in wb.sheetnames:
        summary = wb[summary_sheet]
        for row in range(1, summary.max_row + 1):
            summary_rows.append((
                summary.cell(row=row, column=1).value,  # name
                summary.cell(row=row, column=2).value,   # count
                summary.cell(row=row, column=3).value,   # required
                summary.cell(row=row, column=4).value,   # optional
                summary.cell(row=row, column=5).value,   # total
            ))

    return compute_figures(detail_rows, summary_rows, name_to_key=name_to_key)
