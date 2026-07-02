"""
Single source of truth for estimate figures.

Reads the estimates spreadsheet and computes per-estimate day figures directly
from the raw "Task Detail" rows (the same logic the Summary sheet's SUMIFS
formulas use). Every generator - the walkthroughs, the standalone HTML
presentation, and the PPTX - imports from here rather than hardcoding numbers,
so a single spreadsheet edit + regenerate propagates everywhere.

Workflow:
    1. Edit task days / optional flags in the spreadsheet's "Task Detail" tab
    2. Save
    3. Regenerate (see regenerate_all.py)

Estimate 3a (Training & Enablement) has no task breakdown; it is a manual row
on the Summary sheet and is read from there.

Usage:
    from figures import FIGURES, fmt, effort_line
    FIGURES["est1"].total      # -> 13.5
    fmt(FIGURES["est1"].total) # -> "13.5"
    effort_line("est1")        # -> "~13.5 developer days (9.0 required + testing)"
"""
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

SPREADSHEET = Path(__file__).with_name("BRYT Contract Note Estimates.xlsx")

# Maps the estimate name in the spreadsheet to a stable key used by callers.
_NAME_TO_KEY = {
    "Est 1: PDF/Template Management": "est1",
    "Est 2: DocuSign Integration": "est2",
    "Est 3b: Data Source Extensibility": "est3b",
    "Est 4: Bespoke Contracts": "est4",
    "Est 5: Comparison Audit": "est5",
}
# Estimate 3a is a manual Summary-sheet row (no task breakdown).
_EST_3A_SUMMARY_NAME = "Est 3a: Training & Enablement"


@dataclass(frozen=True)
class Figure:
    key: str
    name: str
    required: float
    optional: float
    total: float
    task_count: int


def fmt(value: float) -> str:
    """Format a day figure: one decimal place, trailing '.0' trimmed to whole days."""
    rounded = round(float(value), 1)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:.1f}"


def _load() -> dict[str, Figure]:
    if not SPREADSHEET.exists():
        raise FileNotFoundError(f"Estimates spreadsheet not found: {SPREADSHEET}")

    wb = load_workbook(SPREADSHEET, data_only=False)
    detail = wb["Task Detail"]

    acc: dict[str, dict] = {}
    for row in range(2, detail.max_row + 1):
        name = detail.cell(row=row, column=1).value
        days = detail.cell(row=row, column=5).value
        optional = detail.cell(row=row, column=6).value
        if not name or days is None or name not in _NAME_TO_KEY:
            continue
        key = _NAME_TO_KEY[name]
        entry = acc.setdefault(key, {"name": name, "required": 0.0, "optional": 0.0, "count": 0})
        entry["count"] += 1
        if optional == "Yes":
            entry["optional"] += float(days)
        else:
            entry["required"] += float(days)

    figures: dict[str, Figure] = {}
    for key, e in acc.items():
        figures[key] = Figure(
            key=key,
            name=e["name"],
            required=round(e["required"], 3),
            optional=round(e["optional"], 3),
            total=round(e["required"] + e["optional"], 3),
            task_count=e["count"],
        )

    # Estimate 3a - manual row on the Summary sheet (columns: name, tasks, required, optional, total)
    summary = wb["Summary"]
    for row in range(1, summary.max_row + 1):
        if summary.cell(row=row, column=1).value == _EST_3A_SUMMARY_NAME:
            req = summary.cell(row=row, column=3).value or 0
            opt = summary.cell(row=row, column=4).value or 0
            total = summary.cell(row=row, column=5).value
            count = summary.cell(row=row, column=2).value or 0
            # Only use if these are literal numbers (not unresolved formula strings)
            if isinstance(req, (int, float)) and isinstance(opt, (int, float)):
                figures["est3a"] = Figure(
                    key="est3a",
                    name=_EST_3A_SUMMARY_NAME,
                    required=float(req),
                    optional=float(opt),
                    total=float(total) if isinstance(total, (int, float)) else float(req) + float(opt),
                    task_count=int(count) if isinstance(count, (int, float)) else 0,
                )
            break

    # Estimate 3 combined (3a + 3b) - convenience for the combined walkthrough
    if "est3a" in figures and "est3b" in figures:
        a, b = figures["est3a"], figures["est3b"]
        figures["est3"] = Figure(
            key="est3",
            name="Est 3: Training & Data Sources",
            required=round(a.required + b.required, 3),
            optional=round(a.optional + b.optional, 3),
            total=round(a.total + b.total, 3),
            task_count=a.task_count + b.task_count,
        )

    return figures


FIGURES: dict[str, Figure] = _load()


def effort_line(key: str) -> str:
    """Cover-badge effort string, e.g. '~13.5 developer days (9.0 required + testing)'."""
    f = FIGURES[key]
    return f"~{fmt(f.total)} developer days ({fmt(f.required)} required + testing)"


def grand_total() -> Figure:
    """Sum across all estimates (1, 2, 3a, 3b, 4, 5)."""
    keys = ["est1", "est2", "est3a", "est3b", "est4", "est5"]
    req = sum(FIGURES[k].required for k in keys if k in FIGURES)
    opt = sum(FIGURES[k].optional for k in keys if k in FIGURES)
    count = sum(FIGURES[k].task_count for k in keys if k in FIGURES)
    return Figure("total", "TOTAL", round(req, 3), round(opt, 3), round(req + opt, 3), count)


if __name__ == "__main__":
    for k in ["est1", "est2", "est3a", "est3b", "est3", "est4", "est5"]:
        if k in FIGURES:
            f = FIGURES[k]
            print(f"{k:6} {f.name:38} req={fmt(f.required):>5}  opt={fmt(f.optional):>4}  total={fmt(f.total):>5}  tasks={f.task_count}")
    gt = grand_total()
    print(f"{'TOTAL':6} {'':38} req={fmt(gt.required):>5}  opt={fmt(gt.optional):>4}  total={fmt(gt.total):>5}  tasks={gt.task_count}")
