import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Task weights per type (in days)
weights = {
    'infrastructure': 1.0,
    'api_backend': 0.5,
    'frontend': 0.75,
    'testing': 0.5,
    'prompt_iteration': 3.0,
    'integration': 0.75,
    'checkpoint': 0,
}

def classify_task(task_text):
    t = task_text.lower()
    if 'checkpoint' in t:
        return 'checkpoint'
    if 'property test' in t or 'integration test' in t or 'unit test' in t:
        return 'testing'
    if 'cdk' in t or 'infrastructure' in t or 'iam' in t or 'trust policy' in t or 'athena workgroup' in t or 'wire cdk' in t or 'configure athena' in t:
        return 'infrastructure'
    if 'angular' in t or 'component' in t or 'frontend' in t or 'module' in t or ('service' in t and 'implement' in t):
        return 'frontend'
    if 'prompt iteration' in t or 'iteration cycle' in t:
        return 'prompt_iteration'
    if 'integration wiring' in t or 'navigation entry' in t or 'wire cdk deployment' in t:
        return 'integration'
    return 'api_backend'

# Read all task files
task_files = [
    ('.kiro/specs/contract-note-template-management/tasks.md', 'Est 1: PDF/Template Management'),
    ('.kiro/specs/contract-note-docusign-integration/tasks.md', 'Est 2: DocuSign Integration'),
    ('.kiro/specs/contract-note-data-source-extensibility/tasks.md', 'Est 3b: Data Source Extensibility'),
    ('.kiro/specs/contract-note-bespoke-contracts/tasks.md', 'Est 4: Bespoke Contracts'),
    ('.kiro/specs/contract-note-comparison-audit/tasks.md', 'Est 5: Comparison Audit'),
]

all_tasks = []
estimate_summaries = []

for filepath, estimate_name in task_files:
    with open(filepath) as f:
        content = f.read()
    
    # Find sub-tasks
    lines = content.split('\n')
    estimate_tasks = []
    
    for line in lines:
        match = re.match(r'\s+- \[ \](\*?)\s+(\d+\.\d+[a-z]?)\s+(.*)', line)
        if match:
            optional_marker = match.group(1)
            task_id = match.group(2)
            task_text = match.group(3).strip()
            is_optional = optional_marker == '*'
            category = classify_task(task_text)
            days = weights[category]
            
            estimate_tasks.append({
                'estimate': estimate_name,
                'task_id': task_id,
                'task': task_text,
                'category': category,
                'days': days,
                'optional': is_optional
            })
    
    total_days = sum(t['days'] for t in estimate_tasks)
    required_days = sum(t['days'] for t in estimate_tasks if not t['optional'])
    optional_days = sum(t['days'] for t in estimate_tasks if t['optional'])
    
    estimate_summaries.append({
        'name': estimate_name,
        'total_tasks': len(estimate_tasks),
        'total_days': total_days,
        'required_days': required_days,
        'optional_days': optional_days,
    })
    
    all_tasks.extend(estimate_tasks)

# Print summary
print("=" * 60)
print("ESTIMATE SUMMARY")
print("=" * 60)
for s in estimate_summaries:
    print(f"{s['name']}")
    print(f"  Tasks: {s['total_tasks']}, Required: {s['required_days']:.1f}d, Optional: {s['optional_days']:.1f}d, Total: {s['total_days']:.1f}d")
print("-" * 60)
total_all = sum(s['total_days'] for s in estimate_summaries)
required_all = sum(s['required_days'] for s in estimate_summaries)
print(f"TOTAL: Required: {required_all:.1f}d, With optional: {total_all:.1f}d")
print("=" * 60)

# Add Estimate 3a (training) manually
estimate_summaries.insert(2, {
    'name': 'Est 3a: Training & Enablement',
    'total_tasks': 7,
    'total_days': 8.0,
    'required_days': 8.0,
    'optional_days': 0,
})

# Create Excel workbook
wb = Workbook()

# Sheet 1: Summary
ws_summary = wb.active
ws_summary.title = "Summary"

header_font = Font(bold=True, size=12)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

ws_summary['A1'] = 'BRYT Contract Note Rework - Estimate Summary'
ws_summary['A1'].font = Font(bold=True, size=14)
ws_summary.merge_cells('A1:E1')

headers = ['Estimate', 'Sub-Tasks', 'Required (days)', 'Optional (days)', 'Total (days)']
for col, h in enumerate(headers, 1):
    cell = ws_summary.cell(row=3, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

for row_idx, s in enumerate(estimate_summaries, 4):
    ws_summary.cell(row=row_idx, column=1, value=s['name']).border = thin_border
    ws_summary.cell(row=row_idx, column=2, value=s['total_tasks']).border = thin_border
    ws_summary.cell(row=row_idx, column=3, value=s['required_days']).border = thin_border
    ws_summary.cell(row=row_idx, column=4, value=s['optional_days']).border = thin_border
    ws_summary.cell(row=row_idx, column=5, value=s['total_days']).border = thin_border

# Totals row
total_row = 4 + len(estimate_summaries)
ws_summary.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
ws_summary.cell(row=total_row, column=2, value=sum(s['total_tasks'] for s in estimate_summaries)).font = Font(bold=True)
ws_summary.cell(row=total_row, column=3, value=sum(s['required_days'] for s in estimate_summaries)).font = Font(bold=True)
ws_summary.cell(row=total_row, column=4, value=sum(s['optional_days'] for s in estimate_summaries)).font = Font(bold=True)
ws_summary.cell(row=total_row, column=5, value=sum(s['total_days'] for s in estimate_summaries)).font = Font(bold=True)
for col in range(1, 6):
    ws_summary.cell(row=total_row, column=col).border = thin_border

ws_summary.column_dimensions['A'].width = 35
ws_summary.column_dimensions['B'].width = 12
ws_summary.column_dimensions['C'].width = 16
ws_summary.column_dimensions['D'].width = 16
ws_summary.column_dimensions['E'].width = 14

# Sheet 2: Task Detail
ws_detail = wb.create_sheet("Task Detail")

detail_headers = ['Estimate', 'Task ID', 'Task Description', 'Category', 'Days', 'Optional']
for col, h in enumerate(detail_headers, 1):
    cell = ws_detail.cell(row=1, column=col, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.border = thin_border

for row_idx, t in enumerate(all_tasks, 2):
    ws_detail.cell(row=row_idx, column=1, value=t['estimate']).border = thin_border
    ws_detail.cell(row=row_idx, column=2, value=t['task_id']).border = thin_border
    ws_detail.cell(row=row_idx, column=3, value=t['task']).border = thin_border
    ws_detail.cell(row=row_idx, column=4, value=t['category']).border = thin_border
    ws_detail.cell(row=row_idx, column=5, value=t['days']).border = thin_border
    ws_detail.cell(row=row_idx, column=6, value='Yes' if t['optional'] else '').border = thin_border

ws_detail.column_dimensions['A'].width = 30
ws_detail.column_dimensions['B'].width = 8
ws_detail.column_dimensions['C'].width = 60
ws_detail.column_dimensions['D'].width = 15
ws_detail.column_dimensions['E'].width = 8
ws_detail.column_dimensions['F'].width = 10

# Save
output_path = 'analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx'
wb.save(output_path)
print(f"\nSpreadsheet saved to: {output_path}")
