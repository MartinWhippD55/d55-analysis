from openpyxl import load_workbook

wb = load_workbook('analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx', data_only=False)
ws = wb['Task Detail']

# Calculate totals per estimate from Task Detail
estimates = {}
order = []
for row in range(2, ws.max_row + 1):
    est = ws.cell(row=row, column=1).value
    days = ws.cell(row=row, column=5).value
    optional = ws.cell(row=row, column=6).value
    if est and days is not None:
        if est not in estimates:
            estimates[est] = {'required': 0, 'optional': 0, 'count': 0}
            order.append(est)
        estimates[est]['count'] += 1
        if optional == 'Yes':
            estimates[est]['optional'] += days
        else:
            estimates[est]['required'] += days

total_req = 0
total_opt = 0
print("=" * 70)
print("UPDATED ESTIMATE FIGURES")
print("=" * 70)
for name in order:
    data = estimates[name]
    total = data['required'] + data['optional']
    total_req += data['required']
    total_opt += data['optional']
    print(f"{name}")
    print(f"  Tasks: {data['count']}, Required: {data['required']}d, Optional: {data['optional']}d, Total: {total}d")

print()
print("Est 3a: Training & Enablement")
print("  Tasks: 7, Required: 8d, Optional: 0d, Total: 8d")
print()
print("-" * 70)
grand_total = total_req + total_opt + 8
print(f"GRAND TOTAL: Required: {total_req + 8}d, Optional: {total_opt}d, Total: {grand_total}d")
print("=" * 70)
