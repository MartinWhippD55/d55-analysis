from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

wb = load_workbook('analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx')

# Delete old summary, create new
if 'Summary' in wb.sheetnames:
    del wb['Summary']

ws = wb.create_sheet('Summary', 0)

header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, size=11, color='FFFFFF')
bold_font = Font(bold=True, size=11)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

ws['A1'] = 'BRYT Contract Note Rework - Estimate Summary'
ws['A1'].font = title_font
ws.merge_cells('A1:E1')

# Headers
headers = ['Estimate', 'Sub-Tasks', 'Required (days)', 'Optional (days)', 'Total (days)']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Estimate names
estimate_names = [
    'Est 1: PDF/Template Management',
    'Est 2: DocuSign Integration',
    'Est 3a: Training & Enablement',
    'Est 3b: Data Source Extensibility',
    'Est 4: Bespoke Contracts',
    'Est 5: Comparison Audit',
]

for row_idx, est_name in enumerate(estimate_names, 4):
    ws.cell(row=row_idx, column=1, value=est_name).border = thin_border

    if est_name == 'Est 3a: Training & Enablement':
        # Manual entry - not in task detail sheet
        ws.cell(row=row_idx, column=2, value=7).border = thin_border
        ws.cell(row=row_idx, column=3, value=8).border = thin_border
        ws.cell(row=row_idx, column=4, value=0).border = thin_border
        ws.cell(row=row_idx, column=5, value=8).border = thin_border
    else:
        # Sub-tasks count: COUNTIF
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.value = f"=COUNTIF('Task Detail'!A:A,A{row_idx})"
        cell_b.border = thin_border

        # Required days: SUMIFS where Optional is blank
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.value = f'=SUMIFS(\'Task Detail\'!E:E,\'Task Detail\'!A:A,A{row_idx},\'Task Detail\'!F:F,"")'
        cell_c.border = thin_border

        # Optional days: SUMIFS where Optional = "Yes"
        cell_d = ws.cell(row=row_idx, column=4)
        cell_d.value = f'=SUMIFS(\'Task Detail\'!E:E,\'Task Detail\'!A:A,A{row_idx},\'Task Detail\'!F:F,"Yes")'
        cell_d.border = thin_border

        # Total = Required + Optional
        cell_e = ws.cell(row=row_idx, column=5)
        cell_e.value = f"=C{row_idx}+D{row_idx}"
        cell_e.border = thin_border

# Totals row
total_row = 4 + len(estimate_names)
ws.cell(row=total_row, column=1, value='TOTAL').font = bold_font
ws.cell(row=total_row, column=1).border = thin_border
for col in range(2, 6):
    cell = ws.cell(row=total_row, column=col)
    col_letter = chr(64 + col)
    cell.value = f"=SUM({col_letter}4:{col_letter}{total_row - 1})"
    cell.font = bold_font
    cell.border = thin_border

ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14

wb.save('analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx')
print('Summary sheet rebuilt with formulas referencing Task Detail')
