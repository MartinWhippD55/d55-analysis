from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

wb = load_workbook('analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx')
ws_detail = wb['Task Detail']

# Get the estimate names in order from Task Detail
estimate_order = []
for row in range(2, ws_detail.max_row + 1):
    val = ws_detail.cell(row=row, column=1).value
    if val and val not in estimate_order:
        estimate_order.append(val)

# Insert Est 3a before Est 3b
est3b_pos = estimate_order.index('Est 3b: Data Source Extensibility')
estimate_order.insert(est3b_pos, 'Est 3a: Training & Enablement')

max_detail_row = ws_detail.max_row

# Rebuild Summary sheet
if 'Summary' in wb.sheetnames:
    del wb['Summary']

ws = wb.create_sheet('Summary', 0)

# Styling
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_font = Font(bold=True, size=11, color='FFFFFF')
bold_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Title
ws['A1'] = 'BRYT Contract Note Rework - Estimate Summary'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:E1')

# Headers
headers = ['Estimate', 'Sub-Tasks', 'Required (days)', 'Optional (days)', 'Total (days)']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border

# Data rows with formulas
for row_idx, est_name in enumerate(estimate_order, 4):
    ws.cell(row=row_idx, column=1, value=est_name).border = thin_border

    if est_name == 'Est 3a: Training & Enablement':
        # Manual row - no detail data for this one
        ws.cell(row=row_idx, column=2, value=7).border = thin_border
        ws.cell(row=row_idx, column=3, value=8).border = thin_border
        ws.cell(row=row_idx, column=4, value=0).border = thin_border
        ws.cell(row=row_idx, column=5, value=8).border = thin_border
    else:
        # Sub-tasks count
        formula_count = f"=COUNTIF('Task Detail'!A:A,A{row_idx})"
        ws.cell(row=row_idx, column=2, value=formula_count).border = thin_border

        # Required days: sum where estimate matches AND optional is blank
        formula_required = (
            f"=SUMPRODUCT(('Task Detail'!A$2:A${max_detail_row}=A{row_idx})"
            f"*('Task Detail'!F$2:F${max_detail_row}=\"\")"
            f"*'Task Detail'!E$2:E${max_detail_row})"
        )
        ws.cell(row=row_idx, column=3, value=formula_required).border = thin_border

        # Optional days: sum where estimate matches AND optional = "Yes"
        formula_optional = (
            f"=SUMPRODUCT(('Task Detail'!A$2:A${max_detail_row}=A{row_idx})"
            f"*('Task Detail'!F$2:F${max_detail_row}=\"Yes\")"
            f"*'Task Detail'!E$2:E${max_detail_row})"
        )
        ws.cell(row=row_idx, column=4, value=formula_optional).border = thin_border

        # Total = Required + Optional
        formula_total = f"=C{row_idx}+D{row_idx}"
        ws.cell(row=row_idx, column=5, value=formula_total).border = thin_border

# Totals row
total_row = 4 + len(estimate_order)
ws.cell(row=total_row, column=1, value='TOTAL').font = bold_font
ws.cell(row=total_row, column=1).border = thin_border
ws.cell(row=total_row, column=2, value=f'=SUM(B4:B{total_row-1})').font = bold_font
ws.cell(row=total_row, column=2).border = thin_border
ws.cell(row=total_row, column=3, value=f'=SUM(C4:C{total_row-1})').font = bold_font
ws.cell(row=total_row, column=3).border = thin_border
ws.cell(row=total_row, column=4, value=f'=SUM(D4:D{total_row-1})').font = bold_font
ws.cell(row=total_row, column=4).border = thin_border
ws.cell(row=total_row, column=5, value=f'=SUM(E4:E{total_row-1})').font = bold_font
ws.cell(row=total_row, column=5).border = thin_border

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14

wb.save('analysis/BRYT/contract-note/BRYT Contract Note Estimates.xlsx')
print('Summary sheet rebuilt with formulas referencing Task Detail.')
