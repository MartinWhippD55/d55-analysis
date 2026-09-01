"""
Single source of truth for Report Builder POC estimate figures.

Reads the POC estimates spreadsheet's authoritative "Task Detail" rows directly
with openpyxl (raw numbers - not Excel's cached formula values, which go stale
when a non-Excel tool edits the file) and computes per-phase day figures. Every
generator (presentation) imports from here rather than hardcoding numbers, so one
spreadsheet edit + regenerate propagates everywhere.

Stripped clone of `../../report-builder/deliverables/figures.py`, pointed at the
POC spreadsheet.

Workflow:
    1. Edit task days / optional flags in the spreadsheet's "Task Detail" tab
    2. Save
    3. Regenerate (see regenerate_all.py)

Usage:
    from figures import FIGURES, fmt, effort_line, grand_total
    FIGURES["phase2"].total      # -> 5.5
    fmt(FIGURES["phase2"].total) # -> "5.5"
"""
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

SPREADSHEET = Path(__file__).with_name("BRYT Report Builder POC Estimates.xlsx")


@dataclass(frozen=True)
class Figure:
    key: str
    name: str
    required: float
    optional: float
    total: float
    task_count: int


def fmt(value: float) -> str:
    """Format a day figure: one decimal, trailing '.0' trimmed to whole days."""
    rounded = round(float(value), 1)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:.1f}"


def _phase_key(estimate_label: str) -> str | None:
    """'Phase 2: Query generation + validation' -> 'phase2'."""
    if not estimate_label or not estimate_label.lower().startswith("phase"):
        return None
    num = estimate_label.split(":", 1)[0].replace("Phase", "").strip()
    return f"phase{num}" if num.isdigit() else None


def _load() -> dict[str, Figure]:
    if not SPREADSHEET.exists():
        raise FileNotFoundError(
            f"POC estimates spreadsheet not found: {SPREADSHEET}\n"
            "Run generate_estimates.py first."
        )

    wb = load_workbook(SPREADSHEET, data_only=False)
    detail = wb["Task Detail"]

    acc: dict[str, dict] = {}
    order: list[str] = []
    for row in range(2, detail.max_row + 1):
        label = detail.cell(row=row, column=1).value
        days = detail.cell(row=row, column=5).value
        optional = detail.cell(row=row, column=6).value
        key = _phase_key(label)
        if key is None or days is None:
            continue
        if key not in acc:
            acc[key] = {"name": label, "required": 0.0, "optional": 0.0, "count": 0}
            order.append(key)
        entry = acc[key]
        entry["count"] += 1
        if optional == "Yes":
            entry["optional"] += float(days)
        else:
            entry["required"] += float(days)

    figures: dict[str, Figure] = {}
    for key in order:
        e = acc[key]
        figures[key] = Figure(
            key=key,
            name=e["name"],
            required=round(e["required"], 3),
            optional=round(e["optional"], 3),
            total=round(e["required"] + e["optional"], 3),
            task_count=e["count"],
        )
    figures["_order"] = order  # type: ignore[assignment]
    return figures


_ALL = _load()
_ORDER: list[str] = _ALL.pop("_order")  # type: ignore[assignment]
FIGURES: dict[str, Figure] = _ALL


def phase_keys() -> list[str]:
    """Ordered list of phase keys as they appear in the spreadsheet."""
    return list(_ORDER)


def effort_line(key: str) -> str:
    """Cover-badge effort string, e.g. '~5.5 developer days (5 required + polish)'."""
    f = FIGURES[key]
    return f"~{fmt(f.total)} developer days ({fmt(f.required)} required + polish)"


def grand_total() -> Figure:
    """Sum across all phases."""
    req = sum(FIGURES[k].required for k in _ORDER)
    opt = sum(FIGURES[k].optional for k in _ORDER)
    count = sum(FIGURES[k].task_count for k in _ORDER)
    return Figure("total", "TOTAL", round(req, 3), round(opt, 3), round(req + opt, 3), count)


if __name__ == "__main__":
    for k in _ORDER:
        f = FIGURES[k]
        print(f"{k:8} {f.name:52} req={fmt(f.required):>5}  opt={fmt(f.optional):>4}  "
              f"total={fmt(f.total):>5}  tasks={f.task_count}")
    gt = grand_total()
    print(f"{'TOTAL':8} {'':52} req={fmt(gt.required):>5}  opt={fmt(gt.optional):>4}  "
          f"total={fmt(gt.total):>5}  tasks={gt.task_count}")
