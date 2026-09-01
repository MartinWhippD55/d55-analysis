"""
Generate the Report Builder POC estimates spreadsheet from the POC spec tasks.

Parses `.kiro/specs/report-builder-poc/tasks.md` (top-level numbered tasks
grouped under `### Phase N — ...` headings), classifies each task, assigns a day
weight, and writes an xlsx with a Task Detail sheet (authoritative, one row per
task) and a Summary sheet (per-phase rollups via live SUMIFS/COUNTIF formulas).

This is a STRIPPED clone of the full-feature estimator
(`../../report-builder/deliverables/generate_estimates.py`). The POC carries no
security spine and no production infrastructure, so the weighting table below is
lighter and the categories differ: there is no `security_core`, no heavy
`infrastructure`, and no formal `testing` suite. The POC is throwaway-quality by
design — these figures are for a demo build, not the production build.

The day figures are INFERRED from the task breakdown - the spec has no day
numbers of its own - so the weighting table below is the main lever a reviewer
will want to adjust. Everything downstream (figures.py, presentation) reads from
the resulting spreadsheet, so a single edit here propagates everywhere.

Usage:
    python analysis/BRYT/report-builder-poc/deliverables/generate_estimates.py
"""
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

REPO = Path(__file__).resolve().parents[4]  # workspace root
TASKS = REPO / ".kiro" / "specs" / "report-builder-poc" / "tasks.md"
OUTPUT = Path(__file__).with_name("BRYT Report Builder POC Estimates.xlsx")

# --- Day-weighting table (days per task, by category) -----------------------
# Editable. These are the primary lever for tuning the estimate. POC weights are
# deliberately lighter than the full feature: no security spine, no prod infra.
WEIGHTS = {
    "scaffold": 1.5,     # lightweight project scaffold (api/ + web/), not the strict prod repo
    "core": 1.5,         # domain types, validateDesign, Query_Generator, round-trip, manifest
    "backend": 1.5,      # simple catalog + reports CRUD against a simple store
    "assistant": 2.5,    # Bedrock Converse loop + mutation tools (the star of the demo)
    "run": 1.5,          # run (direct Athena, no SFN), preview, CSV download
    "frontend": 2.5,     # Angular feature module screens / canvas / client model
    "demo": 1.0,         # seed demo data + end-to-end run-through
}

# Per-task overrides (task id -> days) for known-large items where a single
# top-level task carries much more than the category average.
OVERRIDES = {
    "5": 2.0,    # Query_Generator - design -> SQL, joins, demo scope
    "9": 3.0,    # Assistant Converse loop - the demo centrepiece
    "16": 6.0,   # Screens - the seven screens in one task
}


def classify(task_id: str, text: str) -> str:
    t = text.lower()
    if task_id == "1":
        return "scaffold"
    if any(k in t for k in ("seed a demo", "end-to-end demo", "run-through")):
        return "demo"
    if any(k in t for k in ("assistant", "mutation tools", "validate_query", "explain")):
        return "assistant"
    if any(k in t for k in ("run handler", "csv download", "preview handler")):
        return "run"
    if any(k in t for k in ("flow-canvas", "screens", "client `report_design`", "graph mapping")):
        return "frontend"
    if any(k in t for k in ("catalog service", "reports crud")):
        return "backend"
    if any(k in t for k in ("domain types", "join_manifest", "validatedesign",
                            "query_generator", "serialise", "round-trip")):
        return "core"
    return "core"


def parse_tasks():
    lines = TASKS.read_text(encoding="utf-8").splitlines()
    phase_re = re.compile(r"^### Phase (\d+)\s*[-\u2014]\s*(.*)$")
    task_re = re.compile(r"^- \[[ x~!]\](\*?)\s+(\d+)\.\s+(.*)$")
    tasks = []
    phase_no, phase_name = None, None
    for line in lines:
        m = phase_re.match(line)
        if m:
            phase_no = int(m.group(1))
            phase_name = re.sub(r"\s*\([^)]*\)\s*$", "", m.group(2)).strip()
            continue
        m = task_re.match(line)
        if m and phase_no is not None:
            optional = m.group(1) == "*"
            task_id = m.group(2)
            title = m.group(3).strip().strip("`")
            category = classify(task_id, title)
            days = OVERRIDES.get(task_id, WEIGHTS[category])
            tasks.append({
                "phase_no": phase_no,
                "estimate": f"Phase {phase_no}: {phase_name}",
                "task_id": task_id,
                "task": title,
                "category": category,
                "days": days,
                "optional": optional,
            })
    return tasks


def main():
    tasks = parse_tasks()
    phases = []
    for t in tasks:
        if t["estimate"] not in phases:
            phases.append(t["estimate"])

    header_fill = PatternFill("solid", fgColor="1A0A3E")
    header_font = Font(bold=True, size=11, color="D7ECFA")
    border = Border(*(Side(style="thin"),) * 4)

    wb = Workbook()

    # --- Summary sheet (formulas over Task Detail) --------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "BRYT Report Builder POC - Estimate Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:E1")

    heads = ["Phase", "Tasks", "Required (days)", "Optional (days)", "Total (days)"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    first = 4
    for i, label in enumerate(phases):
        r = first + i
        crit = f"'Task Detail'!$A:$A,$A{r}"
        req = f'=SUMIFS(\'Task Detail\'!$E:$E,{crit},\'Task Detail\'!$F:$F,"")'
        opt = f'=SUMIFS(\'Task Detail\'!$E:$E,{crit},\'Task Detail\'!$F:$F,"Yes")'
        ws.cell(row=r, column=1, value=label).border = border
        ws.cell(row=r, column=2, value=f"=COUNTIF('Task Detail'!$A:$A,$A{r})").border = border
        ws.cell(row=r, column=3, value=req).border = border
        ws.cell(row=r, column=4, value=opt).border = border
        ws.cell(row=r, column=5, value=f"=C{r}+D{r}").border = border

    total_r = first + len(phases)
    ws.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True)
    for col, letter in ((2, "B"), (3, "C"), (4, "D"), (5, "E")):
        cell = ws.cell(row=total_r, column=col,
                       value=f"=SUM({letter}{first}:{letter}{total_r - 1})")
        cell.font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=total_r, column=col).border = border

    for col, w in zip("ABCDE", (42, 10, 16, 16, 14)):
        ws.column_dimensions[col].width = w

    # --- Task Detail sheet (authoritative) ----------------------------------
    wd = wb.create_sheet("Task Detail")
    dheads = ["Estimate", "Task ID", "Task Description", "Category", "Days", "Optional"]
    for c, h in enumerate(dheads, 1):
        cell = wd.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    for i, t in enumerate(tasks, 2):
        wd.cell(row=i, column=1, value=t["estimate"]).border = border
        wd.cell(row=i, column=2, value=t["task_id"]).border = border
        wd.cell(row=i, column=3, value=t["task"]).border = border
        wd.cell(row=i, column=4, value=t["category"]).border = border
        wd.cell(row=i, column=5, value=t["days"]).border = border
        wd.cell(row=i, column=6, value="Yes" if t["optional"] else "").border = border
    for col, w in zip("ABCDEF", (36, 8, 62, 15, 8, 10)):
        wd.column_dimensions[col].width = w

    wb.save(OUTPUT)

    # Console summary
    req = sum(t["days"] for t in tasks if not t["optional"])
    opt = sum(t["days"] for t in tasks if t["optional"])
    print(f"Spreadsheet written: {OUTPUT.name}  ({len(tasks)} tasks)")
    print("=" * 62)
    for label in phases:
        pt = [t for t in tasks if t["estimate"] == label]
        pr = sum(t["days"] for t in pt if not t["optional"])
        po = sum(t["days"] for t in pt if t["optional"])
        print(f"{label:44} req={pr:5.1f} opt={po:4.1f} total={pr + po:5.1f} ({len(pt)})")
    print("-" * 62)
    print(f"{'TOTAL':44} req={req:5.1f} opt={opt:4.1f} total={req + opt:5.1f} ({len(tasks)})")


if __name__ == "__main__":
    main()
