"""Generate AI-DLC workshop spreadsheets (Value Proposition Canvas + Service Catalogue)."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_value_proposition():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI-DLC Value Proposition"

    # Styles
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for i, w in enumerate([5, 40, 5, 40, 5, 40], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("B1:F1")
    ws["B1"] = "AI-DLC Workshop \u2014 Value Proposition Canvas"
    ws["B1"].font = Font(name="Calibri", size=16, bold=True)

    ws.merge_cells("B2:F2")
    ws["B2"] = "D55 AI Development Lifecycle Assessment"
    ws["B2"].font = Font(name="Calibri", size=11, italic=True, color="64748B")

    # Row 4: Vertical / Sub-segment / Prospects
    row = 4
    for col, text in [("B", "Vertical Name"), ("D", "Sub-Segment"), ("F", "Prospective Customers")]:
        ws[f"{col}{row}"] = text
        ws[f"{col}{row}"].font = section_font
        ws[f"{col}{row}"].fill = section_fill

    ws["B5"] = "AI Adoption & Transformation"
    ws["D5"] = "Mid-market organisations (250-2000 employees) looking to adopt or scale AI"
    ws["F5"] = "(List target customers here)"
    for c in ["B5", "D5", "F5"]:
        ws[c].alignment = wrap

    # Section: Pain Points
    row = 7
    for col, text in [("B", "Customer Pain Points"), ("D", "Our Value Proposition"), ("F", "Metrics We Impact")]:
        ws[f"{col}{row}"] = text
        ws[f"{col}{row}"].font = section_font
        ws[f"{col}{row}"].fill = section_fill

    pains = [
        (
            "No clear AI strategy \u2014 experimentation without business alignment",
            "Free AI-DLC Assessment: 1-hour structured workshop producing a radar chart of current vs target maturity across 8 dimensions",
            "Time to AI strategy: weeks \u2192 1 hour\nClarity of AI investment priorities",
        ),
        (
            "Tool sprawl \u2014 too many disconnected platforms, high onboarding time",
            "Platform consolidation advisory + prescriptive tooling workshops for dev teams",
            "Tools touched per week\nNew joiner productivity time\nPlatform maintenance burden",
        ),
        (
            "AI stays in notebooks \u2014 can't ship to production reliably",
            "MLOps implementation + CI/CD for models + observability",
            "Models in production\nDeployment frequency\nMTTR for model failures",
        ),
        (
            "No visibility of AI spend or ROI \u2014 CFO can't justify investment",
            "FinOps programme + per-project cost allocation + ROI measurement framework",
            "Cost per AI project\nROI per initiative\nBudget forecast accuracy",
        ),
        (
            "Governance gap \u2014 no policy, no compliance posture, EU AI Act exposure",
            "AI Governance framework + compliance posture review + region lockdown",
            "Audit readiness time\nPolicy coverage\nRegulatory risk score",
        ),
        (
            "Skills gap \u2014 dependent on heroes, no career path, can't recruit",
            "Embedded squads + training/upskilling + forward deployed engineers",
            "Bus factor\nTime to backfill\nInternal AI literacy rate",
        ),
    ]

    for i, (pain, vp, metrics) in enumerate(pains):
        r = row + 1 + i
        ws[f"B{r}"] = pain
        ws[f"D{r}"] = vp
        ws[f"F{r}"] = metrics
        for c in [f"B{r}", f"D{r}", f"F{r}"]:
            ws[c].alignment = wrap

    # Section: Transformation & Growth
    row = 15
    for col, text in [("B", "Customer Transformation & Growth"), ("D", "Our Value Proposition"), ("F", "Metrics We Impact")]:
        ws[f"{col}{row}"] = text
        ws[f"{col}{row}"].font = section_font
        ws[f"{col}{row}"].fill = section_fill

    transforms = [
        (
            "Want AI to be a competitive differentiator, not just a cost play",
            "AI Strategy Workshop \u2192 Use Case Prioritisation \u2192 Business Case Development",
            "Revenue from AI-enabled products\nCompetitive win rate\nTime to market for AI features",
        ),
        (
            "Need to scale AI from 1-2 use cases to a portfolio",
            "Discovery & Strategy Session (DSD) producing a costed roadmap + runbook",
            "Active AI use cases\nPortfolio ROI\nReuse of platform/data assets",
        ),
        (
            "Want to embed AI literacy across the business, not just tech team",
            "AI Champions Programme + Change Management + Ceremony/Process changes",
            "Non-tech AI tool adoption\nIdeas submitted from frontline\nTraining completion rates",
        ),
    ]

    for i, (growth, vp, metrics) in enumerate(transforms):
        r = row + 1 + i
        ws[f"B{r}"] = growth
        ws[f"D{r}"] = vp
        ws[f"F{r}"] = metrics
        for c in [f"B{r}", f"D{r}", f"F{r}"]:
            ws[c].alignment = wrap

    # Section: Why Us / Why Now
    row = 20
    for col, text in [("B", "WHY D55?"), ("D", "WHY NOW?"), ("F", "TECHNOLOGIES")]:
        ws[f"{col}{row}"] = text
        ws[f"{col}{row}"].font = section_font
        ws[f"{col}{row}"].fill = section_fill

    ws["B21"] = (
        "\u2022 AWS Advanced Partner with AI/ML specialisation\n"
        "\u2022 Proven delivery across data platforms, MLOps, and GenAI\n"
        "\u2022 Forward deployed engineers who build AND upskill\n"
        "\u2022 Free assessment lowers barrier to engagement\n"
        "\u2022 Prescriptive runbooks \u2014 not just advice, but a playbook"
    )
    ws["D21"] = (
        "\u2022 EU AI Act compliance deadlines approaching\n"
        "\u2022 AI moving from experiment to board-level priority\n"
        "\u2022 Competitors adopting \u2014 window to differentiate closing\n"
        "\u2022 Cost pressure demands ROI visibility\n"
        "\u2022 Talent market means build-vs-buy decision is now"
    )
    ws["F21"] = (
        "\u2022 AWS SageMaker Unified Studio\n"
        "\u2022 Bedrock / GenAI\n"
        "\u2022 Glue / Athena / Lakehouse\n"
        "\u2022 MLOps (SageMaker Pipelines, CodePipeline)\n"
        "\u2022 QuickSight\n"
        "\u2022 Infrastructure as Code (CDK/Terraform)\n"
        "\u2022 Kiro / AI-assisted development"
    )
    for c in ["B21", "D21", "F21"]:
        ws[c].alignment = wrap

    # Elevator Pitch
    row = 24
    ws.merge_cells(f"B{row}:F{row}")
    ws[f"B{row}"] = "ELEVATOR PITCH"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill

    ws.merge_cells("B25:F25")
    ws["B25"] = (
        "Book an hour with us. We'll show you exactly where you are with AI, where you want to be, "
        "and what it takes to get there \u2014 for free. You leave with a radar chart, a gap analysis, "
        "and a clear roadmap. No commitment, just clarity. If you want us to help you execute, "
        "we've got workshops, embedded engineers, and runbooks ready to go."
    )
    ws["B25"].alignment = wrap
    ws["B25"].font = Font(name="Calibri", size=11, italic=True)

    wb.save("analysis/D55/ai-dlc/spreadsheets/AI-DLC Value Proposition Canvas.xlsx")
    print("Created: AI-DLC Value Proposition Canvas.xlsx")


def create_service_catalogue():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI-DLC Service Catalogue"

    # Styles
    section_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    subsection_font = Font(name="Calibri", size=10, bold=True)
    section_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")

    # Column widths
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 80

    # Title
    ws.merge_cells("B1:C1")
    ws["B1"] = "AI-DLC Workshop \u2014 Service Catalogue"
    ws["B1"].font = Font(name="Calibri", size=16, bold=True)

    ws.merge_cells("B2:C2")
    ws["B2"] = "D55 AI Development Lifecycle Assessment & Delivery Services"
    ws["B2"].font = Font(name="Calibri", size=11, italic=True, color="64748B")

    # --- SERVICE: AI-DLC Assessment Workshop ---
    row = 4
    ws[f"B{row}"] = "SERVICE NAME"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = "AI-DLC Assessment Workshop"
    ws[f"C{row}"].font = section_font
    ws[f"C{row}"].fill = section_fill

    row = 6
    ws[f"B{row}"] = "SERVICE DESCRIPTION"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "A free, structured 1-hour assessment workshop that scores an organisation across 8 AI maturity "
        "dimensions (Strategy, Data, Tooling, Team, Governance, Delivery, Cost, Culture). "
        "Produces a radar chart showing current state vs target state, a gap analysis with recommended "
        "services, and an indicative runbook. The output drives further paid engagements."
    )
    ws[f"C{row}"].alignment = wrap

    row = 8
    ws[f"B{row}"] = "IDEAL CUSTOMER PROFILE"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "\u2022 Mid-market / SME (250\u20132000 employees)\n"
        "\u2022 Some AI experimentation but struggling to scale or show ROI\n"
        "\u2022 Leadership bought into AI but unclear on strategy or next steps\n"
        "\u2022 May have data platform but AI workloads not in production\n"
        "\u2022 Feeling pressure from competitors, regulation, or board\n"
        "\u2022 On AWS or open to AWS (for service delivery alignment)"
    )
    ws[f"C{row}"].alignment = wrap

    row = 10
    ws[f"B{row}"] = "CUSTOMER CHALLENGES"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "\u2022 No clear AI strategy tied to business outcomes\n"
        "\u2022 Tool and platform sprawl \u2014 high friction, long onboarding\n"
        "\u2022 AI work stuck in experimentation, not reaching production\n"
        "\u2022 No governance framework \u2014 EU AI Act exposure unknown\n"
        "\u2022 Can't quantify AI spend or demonstrate ROI to finance\n"
        "\u2022 Skills concentrated in individuals, no scalable capability\n"
        "\u2022 Culture not ready \u2014 AI seen as tech team's problem, not business-wide"
    )
    ws[f"C{row}"].alignment = wrap

    # Qualifying Questions
    row = 12
    ws[f"B{row}"] = "QUALIFYING QUESTIONS"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = ""
    ws[f"C{row}"].fill = section_fill

    qualifying = [
        ("AI Maturity", [
            "Have you started using AI in any capacity (even experimentation)?",
            "Is there executive sponsorship or budget for AI initiatives?",
            "Do you have data/ML practitioners on staff, or is it outsourced?",
            "Have you tried AI developer tooling (Copilot, Claude, etc.)? What happened?",
        ]),
        ("Ambition & Urgency", [
            "Where does AI sit on the board agenda? Is there a deadline or trigger?",
            "Are competitors doing things with AI that worry you?",
            "Is there a compliance trigger (EU AI Act, sector regulation)?",
            "What's changed in the last 6\u201312 months that makes this conversation worth having now?",
        ]),
        ("Scale & Complexity", [
            "How many people would be involved in or affected by AI adoption?",
            "Do you have heavy coordination roles (POs, BAs, PMs)? How many vs engineers?",
            "How many data sources and systems are in play?",
            "What's your current cloud platform, and is that settled?",
            "What's your release cadence today? Days, weeks, months?",
        ]),
        ("Budget & Decision", [
            "Is there budget allocated for AI strategy or implementation work?",
            "Who makes the decision to engage external support?",
            "What does 'decided' look like for you, and by when?",
            "Would you prefer to buy a strategy/roadmap, or buy people and time?",
        ]),
    ]

    row = 13
    for category, questions in qualifying:
        ws[f"B{row}"] = category
        ws[f"B{row}"].font = subsection_font
        ws[f"C{row}"] = "\n".join(f"\u2022 {q}" for q in questions)
        ws[f"C{row}"].alignment = wrap
        row += 1

    # Discovery Questions
    row += 1
    ws[f"B{row}"] = "DISCOVERY QUESTIONS"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = "(Used during the workshop itself)"
    ws[f"C{row}"].font = section_font
    ws[f"C{row}"].fill = section_fill

    discovery = [
        ("Strategy & Alignment", [
            "Who owns AI in your organisation? Is there an executive sponsor?",
            "Can you point to a document linking AI initiatives to business outcomes?",
            "How are AI projects prioritised? What gets funded and why?",
            "If AI budgets were cut tomorrow, who'd fight for them and with what evidence?",
        ]),
        ("Data Readiness", [
            "Where does your data live today? How many systems for a full picture?",
            "Do you have a data catalogue? Do people actually use it?",
            "How long to produce a clean joined dataset across key entities?",
            "What's your biggest data quality pain point?",
            "AI on bad data produces bad outputs faster \u2014 how confident are you in your data foundation?",
        ]),
        ("Tooling & Platform", [
            "How many tools does a practitioner touch in a typical week?",
            "What's the journey from idea to production? How many hand-offs?",
            "If a new joiner started tomorrow, how long before productive?",
            "What's your biggest frustration with current tooling?",
            "Are your devs using AI-assisted tooling (Copilot, Claude, agentic IDEs)? What's the experience?",
        ]),
        ("Team & Capability", [
            "How many people on data/AI full-time vs second hat?",
            "When a senior leaves, what breaks? How long to backfill?",
            "Who's the bottleneck today? What are they spending their time on?",
            "Is there a training programme for AI? Who's been through it?",
            "If you could add three roles tomorrow, what and why?",
        ]),
        ("Governance & Compliance", [
            "Do you have an AI usage policy? Who enforces it?",
            "How do you document what a model does and what decisions it influences?",
            "Are you aware of EU AI Act implications for your use cases?",
            "If a regulator asked 'show me how this decision was made' \u2014 could you?",
            "Are workloads running in the correct region? Is access locked down?",
        ]),
        ("Delivery & Operations", [
            "How many models in production? How did they get there?",
            "What happens when model performance degrades?",
            "What's the gap between experiments and what's shipping?",
            "'Deploy this by Friday' \u2014 realistic or laughable?",
            "How much time in design/spec vs implementation? Has that changed with AI tooling?",
            "When devs go faster, where does the bottleneck shift \u2014 specs? Reviews? Coordination?",
        ]),
        ("Cost & Value", [
            "Do you know AI/data spend per team or project?",
            "Last time finance asked 'what are we getting?' \u2014 what was the answer?",
            "How do you decide to continue, scale, or kill an AI initiative?",
            "AI tooling costs ~\u00a3150\u2013200/seat/month for ~5x productivity \u2014 have you done this maths for your team?",
        ]),
        ("Culture & Adoption", [
            "How do non-technical staff feel about AI? Excited? Threatened? Indifferent?",
            "Are there AI tools in use outside the data/tech team?",
            "If a frontline employee had an AI idea, who would they tell?",
            "Where is resistance coming from? Which roles or personality types push back?",
            "When devs adopted AI and went faster, did coordination roles (POs/BAs/PMs) keep pace or become a bottleneck?",
            "Has anyone left or been moved because they couldn't adapt?",
        ]),
    ]

    row += 1
    for category, questions in discovery:
        ws[f"B{row}"] = category
        ws[f"B{row}"].font = subsection_font
        ws[f"C{row}"] = "\n".join(f"\u2022 {q}" for q in questions)
        ws[f"C{row}"].alignment = wrap
        row += 1

    # Service Deliverables
    row += 1
    ws[f"B{row}"] = "SERVICE DELIVERABLES"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = ""
    ws[f"C{row}"].fill = section_fill

    row += 1
    ws[f"B{row}"] = "Assessment Output (Free)"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "\u2022 8-dimension radar chart (current vs target state)\n"
        "\u2022 Gap analysis with severity scoring\n"
        "\u2022 Recommended services mapped to each gap\n"
        "\u2022 Indicative runbook (phased by priority)\n"
        "\u2022 Engagement model recommendation"
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "Follow-on Services"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "1. Discovery & Strategy Session (DSD) \u2014 \u00a330\u201350k\n"
        "   Fully costed AI/data strategy + roadmap. Pure consultancy. 4\u20136 weeks.\n\n"
        "2. Prescriptive Workshops \u2014 \u00a35\u201315k per module\n"
        "   Modules: AI Strategy | AI-First Delivery Lifecycle (how to spec, contracts upfront, ceremony changes) | "
        "Developer Tooling | Compliance & Governance | MLOps | Process & Ceremonies | "
        "Data Platform | Cost Optimisation | Four Fears (Culture & Change)\n\n"
        "3. Embedded Team / FDEs \u2014 \u00a315\u201325k per engineer/month\n"
        "   8-week prove-it model:\n"
        "   \u2022 Weeks 1\u20132: Assess codebase, tooling, current AI adoption, identify quick wins\n"
        "   \u2022 Weeks 3\u20136: Embed in a squad, apply AI-first delivery on real work, track metrics from day one\n"
        "   \u2022 Weeks 7\u20138: Demonstrate improvement with numbers, produce playbooks, deliver rollout plan\n"
        "   D55 embeds, proves, then enables. Not a permanent dependency.\n"
        "   Once one squad is working, use it as reference to pull others forward.\n\n"
        "4. Runbook & Asset Delivery \u2014 from \u00a320k (or included with DSD)\n"
        "   Prescriptive step-by-step guide with decision points and options.\n"
        "   The runbook IS the asset \u2014 tells them exactly how to get from A to B."
    )
    ws[f"C{row}"].alignment = wrap

    # Benefits
    row += 2
    ws[f"B{row}"] = "BENEFITS TO CUSTOMER"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = ""
    ws[f"C{row}"].fill = section_fill

    row += 1
    ws[f"C{row}"] = (
        "\u2022 Clarity in 1 hour \u2014 know exactly where you stand and what to do next\n"
        "\u2022 No upfront cost \u2014 assessment is free, you choose what (if anything) to buy\n"
        "\u2022 Visual gap analysis makes the case for investment to leadership/PE/board\n"
        "\u2022 Roadmap is actionable \u2014 phased, prioritised, with clear deliverables\n"
        "\u2022 Flexible engagement \u2014 buy consultancy, workshops, or embedded people\n"
        "\u2022 Knowledge transfer built in \u2014 we upskill as we deliver, not a permanent dependency\n"
        "\u2022 Prescriptive not theoretical \u2014 runbooks with specific steps, not just advice\n"
        "\u2022 AWS-native solutions with cost advantage over Snowflake/Databricks\n"
        "\u2022 Proven economics: ~\u00a3150\u2013200/seat/month AI tooling delivers ~5x developer productivity\n"
        "  For a 50-person team: ~\u00a3120k/year in tooling for output equivalent to 200+ engineers\n"
        "  The real cost is delay, not investment"
    )
    ws[f"C{row}"].alignment = wrap

    # Who needs to be in the room
    row += 2
    ws[f"B{row}"] = "WHO WE NEED IN THE ROOM"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = ""
    ws[f"C{row}"].fill = section_fill

    row += 1
    ws[f"B{row}"] = "For the Assessment"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "\u2022 CTO / VP Engineering / Head of Data (decision maker + strategy perspective)\n"
        "\u2022 Senior data/ML engineer (technical reality check)\n"
        "\u2022 Someone from finance or ops (cost/value perspective)\n"
        "\u2022 Product owner or business stakeholder (adoption/culture lens)\n"
        "3\u20135 people. Enough perspectives to score honestly, few enough to have a real conversation."
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "For Workshops"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "Varies by module:\n"
        "\u2022 AI-First Delivery Lifecycle: Engineering leads, POs, Scrum Masters, BAs (the people whose process changes)\n"
        "\u2022 Developer Tooling: Senior engineers, tech leads (hands-on practitioners)\n"
        "\u2022 Compliance & Governance: CTO, legal/compliance, security, data protection\n"
        "\u2022 Four Fears / Culture: Leadership team, HR, people who'll champion or block adoption\n"
        "\u2022 MLOps: Data engineers, ML engineers, platform/infra team\n"
        "\u2022 Cost Optimisation: Finance, engineering leads, FinOps (if exists)\n\n"
        "Typically 4\u201310 people per workshop. Mix of doers and decision makers.\n"
        "Key learning from experience: include the 'middle layer' (POs/BAs/PMs) early \u2014 "
        "they become the bottleneck when devs go faster and need to adapt too."
    )
    ws[f"C{row}"].alignment = wrap

    # Key Frameworks (from proven delivery experience)
    row += 2
    ws[f"B{row}"] = "KEY FRAMEWORKS & INSIGHTS"
    ws[f"B{row}"].font = section_font
    ws[f"B{row}"].fill = section_fill
    ws[f"C{row}"] = "(Proven in delivery \u2014 used during assessment and workshops)"
    ws[f"C{row}"].font = section_font
    ws[f"C{row}"].fill = section_fill

    row += 1
    ws[f"B{row}"] = "Four Fears Framework"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "Diagnostic tool for AI adoption resistance. Four personality types, four fears:\n\n"
        "\u2022 Controller (tasks/past): 'I can't manage what I can't understand' \u2192 "
        "Reframe: AI gives MORE oversight and observability, not less\n"
        "\u2022 Driver (tasks/future): 'This could go wrong and I'll be accountable' \u2192 "
        "Reframe: AI compresses feedback loops, you fail faster and cheaper\n"
        "\u2022 Stabiliser (people/past): 'My position and authority are under threat' \u2192 "
        "Reframe: People who orchestrate AI become MORE valuable\n"
        "\u2022 Influencer (people/future): 'If AI does the impressive stuff, what's my value?' \u2192 "
        "Reframe: AI handles grunt work, you get more stage\n\n"
        "Application: diagnose which fears dominate, tailor intervention per role/person. "
        "Common pattern: POs/BAs/PMs are typically Controllers + Stabilisers."
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "AI-First Delivery Lifecycle"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "Key insight: when implementation is 5x faster, the bottleneck shifts to clarity of intent.\n\n"
        "The lifecycle shifts toward design-up-front (NOT waterfall):\n"
        "1. Analysis & Design \u2014 more time here than traditional agile. Clear specs.\n"
        "2. Contracts & Integration Points \u2014 understand touch points upfront.\n"
        "3. Implementation \u2014 devs with agentic tooling against well-defined specs.\n"
        "4. Testing \u2014 human-in-the-loop + automated (Playwright MCP for regression).\n"
        "5. Documentation & Handover \u2014 AI-generated assets.\n\n"
        "Garbage specs at 5x speed = garbage 5x faster. Design discipline is the multiplier on the multiplier."
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "Middle Layer Bottleneck"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "When devs go 5x faster with AI tooling, coordination roles can't keep pace:\n"
        "\u2022 POs can't write specs fast enough\n"
        "\u2022 BAs can't break down requirements fast enough\n"
        "\u2022 PMs can't coordinate fast enough\n\n"
        "They become the constraint. Unblocked by:\n"
        "\u2022 Action first, metrics to prove it, regular retros\n"
        "\u2022 1-on-1 training and support for those roles\n"
        "\u2022 AI-assisted spec writing and prioritisation\n"
        "\u2022 Not everyone will come around \u2014 plan for it, budget time for it"
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "The Economics"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "\u2022 AI developer tooling: ~\u00a3150\u2013200/seat/month (Copilot + Claude + IDE)\n"
        "\u2022 Observed productivity gain: ~5x (with belief this improves further)\n"
        "\u2022 For 50 engineers: ~\u00a3120k/year in tooling for output of 200+ engineers\n"
        "\u2022 Even at 20% headcount efficiency, that's millions in savings\n"
        "\u2022 No expensive infrastructure or custom model training required\n"
        "\u2022 The real cost is DELAY, not investment\n\n"
        "PE/CFO framing: AI-native engineering org is a valuation multiplier at exit. "
        "Output per \u00a3 of engineering cost is the metric that matters."
    )
    ws[f"C{row}"].alignment = wrap

    row += 1
    ws[f"B{row}"] = "8-Week Prove-It Model"
    ws[f"B{row}"].font = subsection_font
    ws[f"C{row}"] = (
        "Structured embedded engagement (default shape for FDE work):\n\n"
        "Weeks 1\u20132: ASSESS\n"
        "\u2022 Assess codebase, tooling, current AI adoption\n"
        "\u2022 Understand resistance patterns (Four Fears)\n"
        "\u2022 Identify quick wins and high-friction areas\n\n"
        "Weeks 3\u20136: PROVE\n"
        "\u2022 Embed D55 people in an existing squad (or stand up fresh D55-led squad)\n"
        "\u2022 Apply AI-first delivery on a real piece of work\n"
        "\u2022 Track metrics from day one (PRs, cycle time, deployment frequency)\n\n"
        "Weeks 7\u20138: TRANSLATE\n"
        "\u2022 Demonstrate improvement with hard numbers\n"
        "\u2022 Translate ways of working into reusable assets (guides, patterns, configs)\n"
        "\u2022 Produce rollout plan for other squads\n\n"
        "Then: one squad is the reference point to pull others forward. "
        "Assets and playbooks allow teams to self-serve after D55 steps back."
    )
    ws[f"C{row}"].alignment = wrap

    wb.save("analysis/D55/ai-dlc/spreadsheets/AI-DLC Service Catalogue.xlsx")
    print("Created: AI-DLC Service Catalogue.xlsx")


if __name__ == "__main__":
    create_value_proposition()
    create_service_catalogue()
